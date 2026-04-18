import json
from collections import defaultdict
from datetime import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                               GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00

# ─── Load Data ─────────────────────────────────────────────────────────────
import os as _os
_SCRIPT_DIR = _os.path.dirname(_os.path.abspath(__file__))
_DATA_FILE = _os.path.join(_SCRIPT_DIR, 'data', 'league_6130.json')
with open(_DATA_FILE) as f:
    raw = json.load(f)

team_names = {int(k): v for k, v in raw['team_names'].items()}
# Collect all team IDs from both team_names and games (some teams may be missing from team_names)
_ids_from_games = set()
for _g in raw.get('games', []):
    _a = _g['attributes']
    if _a.get('hteam_id'): _ids_from_games.add(_a['hteam_id'])
    if _a.get('vteam_id'): _ids_from_games.add(_a['vteam_id'])
TEAM_IDS = sorted(set(team_names.keys()) | _ids_from_games)

DIVISIONS = {
    14333:"North", 14334:"North", 14335:"North", 14336:"North", 14337:"North",
    14345:"South", 14346:"South", 14347:"South", 14348:"South", 14349:"South",
    14355:"West",  14356:"West",  14357:"West",  14358:"West",  14359:"West",  14360:"West",
}

def get_division(tid):
    """Resolve division by ID lookup first, then fall back to name prefix."""
    if tid in DIVISIONS:
        return DIVISIONS[tid]
    name = team_names.get(tid, '')
    code = name.split(' ')[0] if name else ''
    if 'N' in code: return 'North'
    if 'S' in code: return 'South'
    if 'W' in code: return 'West'
    return 'Unknown'

RESOURCE_NAMES = {
    1:"Red Rink (Invisalign Arena)", 2:"Black Rink (Invisalign Arena)",
    17:"Wake Rink", 18:"Forest Rink",
    28:"Raleigh Rink", 29:"Cary Rink", 30:"Garner Rink",
    58:"Hillsborough Rink", 59:"Garner Rink"
}

# ─── Parse Games ────────────────────────────────────────────────────────────
all_games = sorted(raw['games'], key=lambda g: g['attributes']['start'])
completed, upcoming = [], []
# gd_map[(team_a, team_b)] = team_a GD
gd_map = {}

for g in all_games:
    a = g['attributes']
    h, v = a['hteam_id'], a['vteam_id']
    if h is None or v is None:
        continue
    if a.get('home_score') is not None and a.get('visiting_score') is not None:
        hs, vs = int(a['home_score']), int(a['visiting_score'])
        gd_map[(h, v)] =  hs - vs
        gd_map[(v, h)] =  vs - hs
        completed.append(g)
    else:
        upcoming.append(g)

# ─── Standings ──────────────────────────────────────────────────────────────
stats = defaultdict(lambda: dict(W=0,L=0,T=0,GF=0,GA=0,GP=0,pts=0,hW=0,hL=0,hT=0,aW=0,aL=0,aT=0))
for g in completed:
    a = g['attributes']
    h, v = a['hteam_id'], a['vteam_id']
    hs, vs = int(a['home_score']), int(a['visiting_score'])
    stats[h]['GP']+=1; stats[v]['GP']+=1
    stats[h]['GF']+=hs; stats[h]['GA']+=vs
    stats[v]['GF']+=vs; stats[v]['GA']+=hs
    if hs > vs:
        stats[h]['W']+=1; stats[v]['L']+=1; stats[h]['pts']+=2
        stats[h]['hW']+=1; stats[v]['aL']+=1
    elif vs > hs:
        stats[v]['W']+=1; stats[h]['L']+=1; stats[v]['pts']+=2
        stats[h]['hL']+=1; stats[v]['aW']+=1
    else:
        stats[h]['T']+=1; stats[v]['T']+=1
        stats[h]['pts']+=1; stats[v]['pts']+=1
        stats[h]['hT']+=1; stats[v]['aT']+=1

# ─── Transitive Inference ────────────────────────────────────────────────────
def get_opponents(tid):
    return [b for (a,b) in gd_map if a == tid]

def predict(team_h, team_a):
    """Returns (weighted_avg_gd, confidence_0_to_1, list_of_path_details)"""
    paths = []

    # 1-hop: common direct opponents
    h_opps = set(get_opponents(team_h))
    a_opps = set(get_opponents(team_a))
    common1 = h_opps & a_opps

    for x in common1:
        gd_hx = gd_map.get((team_h, x))
        gd_ax = gd_map.get((team_a, x))
        if gd_hx is not None and gd_ax is not None:
            pred = gd_hx - gd_ax
            paths.append({
                'pred': pred, 'weight': 1.0, 'hops': 1,
                'desc': f"{team_names[team_h]} vs {team_names[x]} ({gd_hx:+d}gd); "
                        f"{team_names[team_a]} vs {team_names[x]} ({gd_ax:+d}gd)"
            })

    # 2-hop: H→X→Y→A (X opp of H, Y opp of A, X played Y)
    for x in h_opps:
        x_opps = set(get_opponents(x))
        for y in x_opps & a_opps:
            if y in (team_h, team_a):
                continue
            gd_hx = gd_map.get((team_h, x))
            gd_xy = gd_map.get((x, y))
            gd_ay = gd_map.get((team_a, y))
            if all(v is not None for v in [gd_hx, gd_xy, gd_ay]):
                pred = gd_hx + gd_xy - gd_ay
                paths.append({
                    'pred': pred, 'weight': 0.5, 'hops': 2,
                    'desc': f"{team_names[team_h]}→{team_names[x]}({gd_hx:+d})→{team_names[y]}({gd_xy:+d})←{team_names[team_a]}({gd_ay:+d})"
                })

    if not paths:
        return None, 0.0, []

    tw = sum(p['weight'] for p in paths)
    avg = sum(p['pred']*p['weight'] for p in paths) / tw
    # confidence: 1-hop count drives it; cap at 1.0
    one_hop = sum(1 for p in paths if p['hops']==1)
    conf = min(1.0, one_hop/3 + 0.1*sum(1 for p in paths if p['hops']==2)/3)
    return avg, conf, paths

# ─── Excel Styles ─────────────────────────────────────────────────────────
def make_border(style='thin'):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)

def hdr_style(ws, cell_ref, text, bg='1F4E79', fg='FFFFFF', bold=True, center=True, size=11):
    c = ws[cell_ref]
    c.value = text
    c.font = Font(name='Arial', bold=bold, color=fg, size=size)
    c.fill = PatternFill('solid', fgColor=bg)
    c.alignment = Alignment(horizontal='center' if center else 'left',
                             vertical='center', wrap_text=True)
    c.border = make_border()

def data_style(ws, cell_ref, value, bold=False, center=True, bg=None, fg='000000', num_fmt=None, size=10):
    c = ws[cell_ref]
    c.value = value
    c.font = Font(name='Arial', bold=bold, color=fg, size=size)
    c.alignment = Alignment(horizontal='center' if center else 'left', vertical='center')
    c.border = make_border('thin')
    if bg:
        c.fill = PatternFill('solid', fgColor=bg)
    if num_fmt:
        c.number_format = num_fmt

DIV_COLORS = {'North':'DEEAF1','South':'E2EFDA','West':'FFF2CC'}
WIN_BG   = 'C6EFCE'; WIN_FG   = '276221'
LOSS_BG  = 'FFC7CE'; LOSS_FG  = '9C0006'
TIE_BG   = 'FFEB9C'; TIE_FG   = '9C6500'
HEADER_DARK  = '1F4E79'
HEADER_MED   = '2E75B6'
HEADER_LIGHT = 'BDD7EE'
SUBHDR = 'D6E4F0'
UPCOMING_BG = 'F2F2F2'

wb = Workbook()

# ══════════════════════════════════════════════════════════════════════
# SHEET 1: GAME RESULTS
# ══════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Game Results"
ws1.freeze_panes = 'A3'
ws1.row_dimensions[1].height = 30
ws1.row_dimensions[2].height = 20

# Title
ws1.merge_cells('A1:N1')
c = ws1['A1']
c.value = "10U ADV LEAGUE (League 6130) — POLAR ICE — GAME RESULTS"
c.font = Font(name='Arial', bold=True, size=14, color='FFFFFF')
c.fill = PatternFill('solid', fgColor=HEADER_DARK)
c.alignment = Alignment(horizontal='center', vertical='center')

hdrs2 = ['#','Date','Time','Home Team','Division','H Score','A Score','Away Team',
         'Division','Goal Diff','Result','Winner','Venue','Type']
cols2 = [4,12,8,28,8,8,8,28,8,10,10,28,20,16]
for i,(h,w) in enumerate(zip(hdrs2,cols2),1):
    col = get_column_letter(i)
    ws1.column_dimensions[col].width = w
    hdr_style(ws1, f'{col}2', h, bg=HEADER_MED)

game_num = 0
row = 3
week_label = None
for g in sorted(all_games, key=lambda x: x['attributes']['start']):
    a = g['attributes']
    h_id, v_id = a['hteam_id'], a['vteam_id']
    if h_id is None or v_id is None:
        continue

    dt = datetime.fromisoformat(a['start'])
    new_week = dt.strftime('%b %d wk')
    if new_week != week_label:
        week_label = new_week
        # week separator row
        ws1.merge_cells(f'A{row}:N{row}')
        c = ws1[f'A{row}']
        c.value = f"  WEEK OF {dt.strftime('%B %d, %Y').upper()}"
        c.font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor='344E6B')
        c.alignment = Alignment(horizontal='left', vertical='center')
        ws1.row_dimensions[row].height = 18
        row += 1

    game_num += 1
    hs = a.get('home_score')
    vs = a.get('visiting_score')
    is_played = hs is not None and vs is not None

    if is_played:
        hs, vs = int(hs), int(vs)
        gd = hs - vs
        if gd > 0:   result,winner,rbg,rfg = 'HOME WIN',team_names.get(h_id,''),WIN_BG,WIN_FG
        elif gd < 0: result,winner,rbg,rfg = 'AWAY WIN',team_names.get(v_id,''),LOSS_BG,LOSS_FG
        else:        result,winner,rbg,rfg = 'TIE','—',TIE_BG,TIE_FG
        gd_str = f"{gd:+d}"
        row_bg = None
    else:
        hs_disp, vs_disp = '—','—'
        result,winner,rbg,rfg,gd_str = 'UPCOMING','—','F2F2F2','666666','—'
        row_bg = UPCOMING_BG

    div_h = DIVISIONS.get(h_id,'?')
    div_v = DIVISIONS.get(v_id,'?')
    matchup = f"{div_h} Div" if div_h==div_v else f"Cross: {div_h}/{div_v}"
    venue = RESOURCE_NAMES.get(a.get('resource_id'), f"Rink #{a.get('resource_id')}")

    vals = [game_num, dt.strftime('%Y-%m-%d'), dt.strftime('%I:%M %p').lstrip('0'),
            team_names.get(h_id,''), div_h,
            hs if is_played else '—', vs if is_played else '—',
            team_names.get(v_id,''), div_v,
            gd_str if is_played else '—',
            result, winner, venue, matchup]

    for ci, val in enumerate(vals, 1):
        col = get_column_letter(ci)
        ref = f'{col}{row}'
        is_num = ci in (1,6,7)
        data_style(ws1, ref, val,
                   center=(ci not in (4,8,12,13)),
                   bg=row_bg if not is_played else None,
                   num_fmt='#,##0' if is_num and is_played else None)
        if ci in (11,):  # Result col
            ws1[ref].fill = PatternFill('solid', fgColor=rbg)
            ws1[ref].font = Font(name='Arial', bold=True, color=rfg, size=10)
        if ci == 10 and is_played:  # GD
            gd_int = int(a['home_score']) - int(a['visiting_score'])
            ws1[ref].font = Font(name='Arial', bold=True,
                                  color=WIN_FG if gd_int>0 else (LOSS_FG if gd_int<0 else TIE_FG), size=10)
        if ci in (4,8) and not is_played:
            ws1[ref].font = Font(name='Arial', color='888888', size=10)

    ws1.row_dimensions[row].height = 16
    row += 1

# ── May weekends — matchups not yet published ──────────────────────────
for tbd_date, tbd_week in [("2026-05-10","Week 7"),("2026-05-17","Week 8"),("2026-05-31","Week 9 (Final)")]:
    dt_tbd = datetime.fromisoformat(tbd_date)
    ws1.merge_cells(f'A{row}:N{row}')
    c = ws1[f'A{row}']
    c.value = f"  WEEK OF {dt_tbd.strftime('%B %d, %Y').upper()}  —  {tbd_week.upper()}  —  MATCHUPS NOT YET PUBLISHED"
    c.font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
    c.fill = PatternFill('solid', fgColor='344E6B')
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws1.row_dimensions[row].height = 18
    row += 1
    for slot in range(1, 9):
        game_num += 1
        vals = [game_num, dt_tbd.strftime('%Y-%m-%d'), 'TBD',
                'Matchup TBD', '?', '—', '—', 'Matchup TBD', '?',
                '—', 'SCHEDULED', '—', 'TBD', '—']
        for ci, val in enumerate(vals, 1):
            ref = f'{get_column_letter(ci)}{row}'
            c = ws1[ref]
            c.value = val
            c.font = Font(name='Arial', color='888888', size=10, italic=True)
            c.fill = PatternFill('solid', fgColor='F0F0F0')
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = make_border('thin')
        ws1.row_dimensions[row].height = 15
        row += 1

# ══════════════════════════════════════════════════════════════════════
# SHEET 2: STANDINGS
# ══════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Standings")
ws2.freeze_panes = 'A4'

ws2.merge_cells('A1:P1')
c = ws2['A1']
c.value = "10U ADV LEAGUE — STANDINGS (as of April 13, 2026)"
c.font = Font(name='Arial', bold=True, size=14, color='FFFFFF')
c.fill = PatternFill('solid', fgColor=HEADER_DARK)
c.alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[1].height = 30

hdrs_st = ['Rank','Team','Div','GP','W','L','T','PTS','GF','GA','GD',
           'W%','Avg GF','Avg GA','Home','Away']
widths_st = [6,30,8,5,5,5,5,6,6,6,6,8,8,8,10,10]
for i,(h,w) in enumerate(zip(hdrs_st,widths_st),1):
    col = get_column_letter(i)
    ws2.column_dimensions[col].width = w
    ws2.row_dimensions[2].height = 8
    hdr_style(ws2, f'{col}3', h, bg=HEADER_MED)

row = 4
div_order = ['North','South','West']
for div in div_order:
    # div header
    ws2.merge_cells(f'A{row}:P{row}')
    c = ws2[f'A{row}']
    c.value = f"  {div.upper()} DIVISION"
    c.font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    c.fill = PatternFill('solid', fgColor='2E75B6')
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws2.row_dimensions[row].height = 18
    row += 1

    div_teams = [t for t in TEAM_IDS if DIVISIONS.get(t)==div]
    div_teams_sorted = sorted(div_teams,
        key=lambda t: (-stats[t]['pts'], -(stats[t]['GF']-stats[t]['GA']), -stats[t]['GF']))

    for rank, tid in enumerate(div_teams_sorted, 1):
        s = stats[tid]
        gd = s['GF'] - s['GA']
        wp = s['W']/s['GP'] if s['GP'] else 0
        avg_gf = s['GF']/s['GP'] if s['GP'] else 0
        avg_ga = s['GA']/s['GP'] if s['GP'] else 0
        home_rec = f"{s['hW']}-{s['hL']}-{s['hT']}"
        away_rec = f"{s['aW']}-{s['aL']}-{s['aT']}"
        bg = DIV_COLORS.get(div,'FFFFFF')

        row_vals = [rank, team_names[tid], div,
                    s['GP'], s['W'], s['L'], s['T'], s['pts'],
                    s['GF'], s['GA'], gd,
                    wp, avg_gf, avg_ga, home_rec, away_rec]

        for ci, val in enumerate(row_vals, 1):
            col = get_column_letter(ci)
            ref = f'{col}{row}'
            is_pct = ci==12
            is_dec = ci in (13,14)
            data_style(ws2, ref, val,
                       center=(ci not in (2,)),
                       bg=bg,
                       num_fmt='0.0%' if is_pct else ('0.0' if is_dec else None),
                       bold=(ci==8))
            if ci==11:  # GD
                ws2[ref].font = Font(name='Arial', bold=True,
                    color=WIN_FG if gd>0 else (LOSS_FG if gd<0 else TIE_FG), size=10)
        ws2.row_dimensions[row].height = 16
        row += 1

row += 1
# Overall standings header
ws2.merge_cells(f'A{row}:P{row}')
c = ws2[f'A{row}']
c.value = "  OVERALL STANDINGS"
c.font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
c.fill = PatternFill('solid', fgColor=HEADER_DARK)
c.alignment = Alignment(horizontal='left', vertical='center')
ws2.row_dimensions[row].height = 18
row += 1

all_sorted = sorted(TEAM_IDS,
    key=lambda t: (-stats[t]['pts'], -(stats[t]['GF']-stats[t]['GA']), -stats[t]['GF']))
for rank, tid in enumerate(all_sorted, 1):
    s = stats[tid]
    gd = s['GF'] - s['GA']
    wp = s['W']/s['GP'] if s['GP'] else 0
    avg_gf = s['GF']/s['GP'] if s['GP'] else 0
    avg_ga = s['GA']/s['GP'] if s['GP'] else 0
    home_rec = f"{s['hW']}-{s['hL']}-{s['hT']}"
    away_rec = f"{s['aW']}-{s['aL']}-{s['aT']}"
    div = DIVISIONS.get(tid,'?')
    bg = DIV_COLORS.get(div,'FFFFFF')
    row_vals = [rank, team_names[tid], div,
                s['GP'], s['W'], s['L'], s['T'], s['pts'],
                s['GF'], s['GA'], gd,
                wp, avg_gf, avg_ga, home_rec, away_rec]
    for ci, val in enumerate(row_vals, 1):
        col = get_column_letter(ci)
        ref = f'{col}{row}'
        is_pct = ci==12; is_dec = ci in (13,14)
        data_style(ws2, ref, val, center=(ci not in (2,)), bg=bg,
                   num_fmt='0.0%' if is_pct else ('0.0' if is_dec else None),
                   bold=(ci in (1,8)))
        if ci==11:
            ws2[ref].font = Font(name='Arial', bold=True,
                color=WIN_FG if gd>0 else (LOSS_FG if gd<0 else TIE_FG), size=10)
    ws2.row_dimensions[row].height = 16
    row += 1

# ══════════════════════════════════════════════════════════════════════
# SHEET 3: PREDICTIONS
# ══════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Predictions")
ws3.freeze_panes = 'A4'

ws3.merge_cells('A1:M1')
c = ws3['A1']
c.value = "10U ADV LEAGUE — UPCOMING GAME PREDICTIONS (Transitive Inference)"
c.font = Font(name='Arial', bold=True, size=14, color='FFFFFF')
c.fill = PatternFill('solid', fgColor=HEADER_DARK)
c.alignment = Alignment(horizontal='center', vertical='center')
ws3.row_dimensions[1].height = 30

ws3.merge_cells('A2:M2')
note = ws3['A2']
note.value = ("Prediction method: Transitive goal differentials.  "
              "1-hop = direct common opponents (weight 1.0).  "
              "2-hop = shared opponents 2 steps away (weight 0.5).  "
              "Confidence = f(# paths found).  Positive predicted GD favors the Home team.")
note.font = Font(name='Arial', italic=True, size=9, color='444444')
note.fill = PatternFill('solid', fgColor='EEF3F9')
note.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
ws3.row_dimensions[2].height = 28

hdrs_p = ['#','Date','Time','Home Team','Division','Away Team','Division',
          'Paths\n(1-hop)','Paths\n(2-hop)','Predicted GD\n(Home)','Confidence',
          'Predicted Winner','Key Inference Paths']
widths_p = [5,12,8,28,8,28,8,9,9,14,12,28,80]
for i,(h,w) in enumerate(zip(hdrs_p,widths_p),1):
    col = get_column_letter(i)
    ws3.column_dimensions[col].width = w
    hdr_style(ws3, f'{col}3', h, bg=HEADER_MED)
ws3.row_dimensions[3].height = 30

row = 4
for gnum, g in enumerate(sorted(upcoming, key=lambda x: x['attributes']['start']), 1):
    a = g['attributes']
    h_id, v_id = a['hteam_id'], a['vteam_id']
    if h_id is None or v_id is None:
        continue
    dt = datetime.fromisoformat(a['start'])
    pred_gd, conf, paths = predict(h_id, v_id)
    hops1 = sum(1 for p in paths if p['hops']==1)
    hops2 = sum(1 for p in paths if p['hops']==2)
    div_h = DIVISIONS.get(h_id,'?')
    div_v = DIVISIONS.get(v_id,'?')

    if pred_gd is not None:
        if pred_gd > 0.5:
            winner = team_names[h_id]; wbg = WIN_BG; wfg = WIN_FG
        elif pred_gd < -0.5:
            winner = team_names[v_id]; wbg = LOSS_BG; wfg = LOSS_FG
        else:
            winner = "TOSS-UP"; wbg = TIE_BG; wfg = TIE_FG
        gd_display = round(pred_gd, 1)
        conf_pct = conf
        # Top 3 path descriptions
        top_paths = sorted(paths, key=lambda p: -p['weight'])[:3]
        path_str = ' | '.join(p['desc'] for p in top_paths)
    else:
        winner = "INSUFFICIENT DATA"; wbg = 'EEEEEE'; wfg = '666666'
        gd_display = None; conf_pct = 0; path_str = "No common opponents yet"

    row_vals = [gnum, dt.strftime('%Y-%m-%d'), dt.strftime('%I:%M %p').lstrip('0'),
                team_names[h_id], div_h, team_names[v_id], div_v,
                hops1, hops2,
                gd_display, conf_pct if pred_gd is not None else None,
                winner, path_str]

    for ci, val in enumerate(row_vals, 1):
        col = get_column_letter(ci)
        ref = f'{col}{row}'
        center = ci not in (4,6,12,13)
        is_pct = ci==11
        num_fmt = '0.0%' if is_pct else ('0.0' if ci==10 and val is not None else None)
        data_style(ws3, ref, val, center=center, num_fmt=num_fmt)
        if ci == 10 and val is not None:  # Pred GD
            fg = WIN_FG if val>0.5 else (LOSS_FG if val<-0.5 else TIE_FG)
            ws3[ref].font = Font(name='Arial', bold=True, color=fg, size=10)
        if ci == 12:  # Winner
            ws3[ref].fill = PatternFill('solid', fgColor=wbg)
            ws3[ref].font = Font(name='Arial', bold=True, color=wfg, size=10)
        if ci == 13:  # Path desc
            ws3[ref].alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
            ws3[ref].font = Font(name='Arial', size=8, color='444444')
    ws3.row_dimensions[row].height = 16
    row += 1

# ── May weekends — matchups not yet published ──────────────────────────
pred_gnum = sum(1 for g in upcoming if g['attributes']['hteam_id'] is not None)
for tbd_date, tbd_week in [("2026-05-10","Week 7"),("2026-05-17","Week 8"),("2026-05-31","Week 9 (Final)")]:
    dt_tbd = datetime.fromisoformat(tbd_date)
    ws3.merge_cells(f'A{row}:M{row}')
    c = ws3[f'A{row}']
    c.value = f"  {tbd_week.upper()}  —  {dt_tbd.strftime('%B %d, %Y').upper()}  —  MATCHUPS NOT YET PUBLISHED  —  8 GAMES SCHEDULED"
    c.font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
    c.fill = PatternFill('solid', fgColor='344E6B')
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws3.row_dimensions[row].height = 18
    row += 1
    for slot in range(1, 9):
        pred_gnum += 1
        tbd_vals = [pred_gnum, dt_tbd.strftime('%Y-%m-%d'), 'TBD',
                    'TBD', '?', 'TBD', '?',
                    '—', '—', '—', '—', 'MATCHUP TBD', 'Schedule not yet published']
        for ci, val in enumerate(tbd_vals, 1):
            ref = f'{get_column_letter(ci)}{row}'
            c = ws3[ref]
            c.value = val
            c.font = Font(name='Arial', color='888888', size=10, italic=True)
            c.fill = PatternFill('solid', fgColor='F0F0F0')
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = make_border('thin')
        ws3.row_dimensions[row].height = 15
        row += 1

# ══════════════════════════════════════════════════════════════════════
# SHEET 4: STRENGTH MATRIX
# ══════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Strength Matrix")

ws4.merge_cells('A1:R1')
c = ws4['A1']
c.value = ("10U ADV LEAGUE — PREDICTED GOAL DIFFERENTIAL MATRIX  "
           "(row team as HOME vs column team as AWAY;  ✓ = actual result,  ~ = predicted)")
c.font = Font(name='Arial', bold=True, size=12, color='FFFFFF')
c.fill = PatternFill('solid', fgColor=HEADER_DARK)
c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
ws4.row_dimensions[1].height = 36

# header row = team names (columns)
sorted_teams = sorted(TEAM_IDS,
    key=lambda t: (-stats[t]['pts'], -(stats[t]['GF']-stats[t]['GA'])))

ws4.column_dimensions['A'].width = 2
ws4.column_dimensions['B'].width = 30
for i, tid in enumerate(sorted_teams, 3):
    col = get_column_letter(i)
    ws4.column_dimensions[col].width = 14
    c = ws4[f'{col}2']
    short = team_names[tid].split(' - ')[-1] if ' - ' in team_names[tid] else team_names[tid]
    c.value = short
    c.font = Font(name='Arial', bold=True, size=8, color='FFFFFF')
    c.fill = PatternFill('solid', fgColor=HEADER_MED)
    c.alignment = Alignment(horizontal='center', vertical='center',
                             text_rotation=45, wrap_text=False)
    c.border = make_border()
ws4.row_dimensions[2].height = 70

# rank header
c = ws4['A2']
c.value = '#'
c.font = Font(name='Arial', bold=True, size=9, color='FFFFFF')
c.fill = PatternFill('solid', fgColor=HEADER_MED)
c.alignment = Alignment(horizontal='center', vertical='center')
c.border = make_border()

c = ws4['B2']
c.value = 'Team (HOME →)'
c.font = Font(name='Arial', bold=True, size=9, color='FFFFFF')
c.fill = PatternFill('solid', fgColor=HEADER_MED)
c.alignment = Alignment(horizontal='center', vertical='center')
c.border = make_border()

for ri, h_id in enumerate(sorted_teams, 3):
    # row label
    rank = ri - 2
    div = DIVISIONS.get(h_id,'?')
    div_bg = DIV_COLORS.get(div,'FFFFFF')
    c = ws4[f'A{ri}']
    c.value = rank
    c.font = Font(name='Arial', bold=True, size=9)
    c.fill = PatternFill('solid', fgColor=div_bg)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = make_border()

    c = ws4[f'B{ri}']
    c.value = team_names[h_id]
    c.font = Font(name='Arial', bold=True, size=9)
    c.fill = PatternFill('solid', fgColor=div_bg)
    c.alignment = Alignment(horizontal='left', vertical='center')
    c.border = make_border()

    for ci, v_id in enumerate(sorted_teams, 3):
        col = get_column_letter(ci)
        ref = f'{col}{ri}'
        if h_id == v_id:
            c = ws4[ref]
            c.value = '—'
            c.fill = PatternFill('solid', fgColor='CCCCCC')
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = make_border()
            continue

        actual = gd_map.get((h_id, v_id))
        if actual is not None:
            display = f"✓ {actual:+d}"
            if actual > 0: bg,fg = WIN_BG, WIN_FG
            elif actual < 0: bg,fg = LOSS_BG, LOSS_FG
            else: bg,fg = TIE_BG, TIE_FG
            bold = True
        else:
            pred_gd, conf, _ = predict(h_id, v_id)
            if pred_gd is not None:
                display = f"~ {pred_gd:+.1f}"
                if pred_gd > 0.5: bg,fg = 'C6EFCE','276221'
                elif pred_gd < -0.5: bg,fg = 'FFD7DC','9C0006'
                else: bg,fg = 'FFEB9C','9C6500'
                bold = False
            else:
                display = "?"
                bg,fg = 'F2F2F2','AAAAAA'
                bold = False

        c = ws4[ref]
        c.value = display
        c.font = Font(name='Arial', bold=bold, color=fg, size=9)
        c.fill = PatternFill('solid', fgColor=bg)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = make_border()
    ws4.row_dimensions[ri].height = 18

# ══════════════════════════════════════════════════════════════════════
# SHEET 5: TEAM GAME LOGS
# ══════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Team Game Logs")
ws5.freeze_panes = 'A3'

ws5.merge_cells('A1:J1')
c = ws5['A1']
c.value = "10U ADV LEAGUE — INDIVIDUAL TEAM GAME LOGS"
c.font = Font(name='Arial', bold=True, size=14, color='FFFFFF')
c.fill = PatternFill('solid', fgColor=HEADER_DARK)
c.alignment = Alignment(horizontal='center', vertical='center')
ws5.row_dimensions[1].height = 30

hdrs5 = ['Team','Division','Date','H/A','Opponent','GF','GA','GD','Result','Running Pts']
widths5 = [28,8,12,5,30,6,6,6,8,12]
for i,(h,w) in enumerate(zip(hdrs5,widths5),1):
    col = get_column_letter(i)
    ws5.column_dimensions[col].width = w
    hdr_style(ws5, f'{col}2', h, bg=HEADER_MED)
ws5.row_dimensions[2].height = 20

row = 3
for tid in sorted_teams:
    div = DIVISIONS.get(tid,'?')
    div_bg = DIV_COLORS.get(div,'FFFFFF')
    team_game_rows = []
    for g in sorted(all_games, key=lambda x: x['attributes']['start']):
        a = g['attributes']
        h_id, v_id = a['hteam_id'], a['vteam_id']
        if h_id is None or v_id is None or tid not in (h_id, v_id):
            continue
        is_home = tid == h_id
        opp_id = v_id if is_home else h_id
        if a.get('home_score') is not None:
            if is_home:
                gf, ga = int(a['home_score']), int(a['visiting_score'])
            else:
                gf, ga = int(a['visiting_score']), int(a['home_score'])
            gd = gf - ga
            if gd > 0: res,rbg,rfg = 'W ✅',WIN_BG,WIN_FG
            elif gd < 0: res,rbg,rfg = 'L ❌',LOSS_BG,LOSS_FG
            else: res,rbg,rfg = 'T 🤝',TIE_BG,TIE_FG
            team_game_rows.append((a['start'][:10],'H' if is_home else 'A',
                                   team_names.get(opp_id,''),gf,ga,gd,res,rbg,rfg))
        else:
            team_game_rows.append((a['start'][:10],'H' if is_home else 'A',
                                   team_names.get(opp_id,''),None,None,None,'⏳','F2F2F2','888888'))

    # Add TBD placeholder rows for unscheduled May weekends
    for tbd_date, tbd_week in [("2026-05-10","Wk 7"),("2026-05-17","Wk 8"),("2026-05-31","Wk 9")]:
        team_game_rows.append((tbd_date, '?', f'TBD — Matchup Not Published ({tbd_week})',
                               None, None, None, '📅 TBD', 'ECECEC', '888888'))

    if not team_game_rows:
        continue

    # team header row
    ws5.merge_cells(f'A{row}:J{row}')
    s = stats[tid]
    gd_tot = s['GF']-s['GA']
    c = ws5[f'A{row}']
    c.value = (f"  {team_names[tid]}  |  {div} Division  |  "
               f"{s['W']}-{s['L']}-{s['T']}  ({s['pts']} pts)  |  "
               f"GF: {s['GF']}  GA: {s['GA']}  GD: {gd_tot:+d}")
    c.font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
    c.fill = PatternFill('solid', fgColor='344E6B')
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws5.row_dimensions[row].height = 18
    row += 1

    running_pts = 0
    for game_date, ha, opp, gf, ga, gd, res, rbg, rfg in team_game_rows:
        if gf is not None:
            if 'W' in res: running_pts += 2
            elif 'T' in res: running_pts += 1

        vals5 = [team_names[tid], div, game_date, ha, opp,
                 gf if gf is not None else '—',
                 ga if ga is not None else '—',
                 f"{gd:+d}" if gd is not None else '—',
                 res,
                 running_pts if gf is not None else '—']

        for ci, val in enumerate(vals5, 1):
            col = get_column_letter(ci)
            ref = f'{col}{row}'
            center = ci not in (1,5,9)
            data_style(ws5, ref, val, center=center,
                       bg=div_bg if gf is not None else UPCOMING_BG)
            if ci == 9:
                ws5[ref].fill = PatternFill('solid', fgColor=rbg)
                ws5[ref].font = Font(name='Arial', bold=True, color=rfg, size=10)
                ws5[ref].alignment = Alignment(horizontal='center', vertical='center')
            if ci == 8 and gd is not None:
                ws5[ref].font = Font(name='Arial', bold=True,
                    color=WIN_FG if gd>0 else (LOSS_FG if gd<0 else TIE_FG), size=10)
        ws5.row_dimensions[row].height = 15
        row += 1
    row += 1  # blank between teams

# ══════════════════════════════════════════════════════════════════════
# SHEET 6: 🥒 DISCO PICKLES SPOTLIGHT
# ══════════════════════════════════════════════════════════════════════
DP_ID       = 14356
DP_DARK     = '145214'
DP_MED      = '1D7A1D'
DP_LIGHT    = 'AAFFAA'
DP_BG       = 'E8FFE8'

ws6 = wb.create_sheet("🥒 Disco Pickles")
ws6.freeze_panes = 'A4'

dp_s  = stats[DP_ID]
dp_gd_tot = dp_s['GF'] - dp_s['GA']

# Column widths
for i, w in enumerate([5,12,8,30,8,8,8,30,10,30,12], 1):
    ws6.column_dimensions[get_column_letter(i)].width = w

# ── Title banner ──────────────────────────────────────────────────────
ws6.merge_cells('A1:K1')
c = ws6['A1']
c.value = (f"🥒  DISCO PICKLES  (10W2)   "
           f"{dp_s['W']}-{dp_s['L']}-{dp_s['T']}  |  {dp_s['pts']} PTS  |  "
           f"GF: {dp_s['GF']}   GA: {dp_s['GA']}   GD: {dp_gd_tot:+d}   |   WEST DIVISION")
c.font = Font(name='Arial', bold=True, size=14, color='FFFFFF')
c.fill = PatternFill('solid', fgColor=DP_DARK)
c.alignment = Alignment(horizontal='center', vertical='center')
ws6.row_dimensions[1].height = 32

def dp_hdr_row(ws, row, label, bg=DP_MED):
    ws.merge_cells(f'A{row}:K{row}')
    c = ws[f'A{row}']
    c.value = f"  {label}"
    c.font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    c.fill = PatternFill('solid', fgColor=bg)
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[row].height = 20

col_hdrs = ['#','Date','Time','Home Team','Div','H Sc','A Sc','Away Team','GD','Result / Prediction','Pts']

def dp_col_headers(ws, row):
    for i, h in enumerate(col_hdrs, 1):
        hdr_style(ws, f'{get_column_letter(i)}{row}', h, bg='2E7D32', fg='FFFFFF')
    ws.row_dimensions[row].height = 18

# ── Section 1: Completed Games ────────────────────────────────────────
row = 2
dp_hdr_row(ws6, row, "✅  COMPLETED GAMES"); row += 1
dp_col_headers(ws6, row); row += 1

running_pts = 0
dp_game_num = 0
dp_played = sorted([g for g in completed
                    if g['attributes']['hteam_id'] == DP_ID
                    or g['attributes']['vteam_id'] == DP_ID],
                   key=lambda x: x['attributes']['start'])

for g in dp_played:
    a = g['attributes']
    dp_game_num += 1
    h_id, v_id = a['hteam_id'], a['vteam_id']
    is_home = (h_id == DP_ID)
    hs, vs = int(a['home_score']), int(a['visiting_score'])
    dp_gf = hs if is_home else vs
    dp_ga = vs if is_home else hs
    gd = dp_gf - dp_ga
    if gd > 0:   res, rbg, rfg = 'WIN ✅',  WIN_BG,  WIN_FG;  running_pts += 2
    elif gd < 0: res, rbg, rfg = 'LOSS ❌', LOSS_BG, LOSS_FG
    else:        res, rbg, rfg = 'TIE 🤝',  TIE_BG,  TIE_FG;  running_pts += 1
    dt = datetime.fromisoformat(a['start'])
    vals = [dp_game_num, dt.strftime('%Y-%m-%d'), dt.strftime('%I:%M %p').lstrip('0'),
            team_names.get(h_id,''), DIVISIONS.get(h_id,'?'),
            hs, vs, team_names.get(v_id,''), f"{gd:+d}", res, running_pts]
    for ci, val in enumerate(vals, 1):
        ref = f'{get_column_letter(ci)}{row}'
        data_style(ws6, ref, val, center=(ci not in (4,8,10)), bg=DP_BG)
        if ci == 9:
            ws6[ref].font = Font(name='Arial', bold=True,
                color=WIN_FG if gd>0 else (LOSS_FG if gd<0 else TIE_FG), size=10)
        if ci == 10:
            ws6[ref].fill = PatternFill('solid', fgColor=rbg)
            ws6[ref].font = Font(name='Arial', bold=True, color=rfg, size=10)
    ws6.row_dimensions[row].height = 16
    row += 1

row += 1  # spacer

# ── Section 2: Upcoming Games (Apr 19 & Apr 26, matchups known) ───────
dp_hdr_row(ws6, row, "📅  UPCOMING GAMES  —  Weeks 5 & 6 (matchups confirmed)"); row += 1
dp_col_headers(ws6, row); row += 1

dp_upcoming_games = sorted([g for g in upcoming
                             if g['attributes']['hteam_id'] == DP_ID
                             or g['attributes']['vteam_id'] == DP_ID],
                            key=lambda x: x['attributes']['start'])

for g in dp_upcoming_games:
    a = g['attributes']
    dp_game_num += 1
    h_id, v_id = a['hteam_id'], a['vteam_id']
    is_home = (h_id == DP_ID)
    pred_gd, conf, paths = predict(h_id, v_id)
    dp_pred = pred_gd if is_home else (-pred_gd if pred_gd is not None else None)
    if dp_pred is not None:
        if dp_pred > 0.5:   pred_txt = f"Predicted WIN  (+{dp_pred:.1f} GD,  {conf*100:.0f}% conf)"
        elif dp_pred < -0.5: pred_txt = f"Predicted LOSS  ({dp_pred:.1f} GD,  {conf*100:.0f}% conf)"
        else:                pred_txt = f"TOSS-UP  ({dp_pred:+.1f} GD,  {conf*100:.0f}% conf)"
        pbg = WIN_BG if dp_pred > 0.5 else (LOSS_BG if dp_pred < -0.5 else TIE_BG)
        pfg = WIN_FG if dp_pred > 0.5 else (LOSS_FG if dp_pred < -0.5 else TIE_FG)
    else:
        pred_txt = "Insufficient data"; pbg = 'EEEEEE'; pfg = '666666'
    dt = datetime.fromisoformat(a['start'])
    vals = [dp_game_num, dt.strftime('%Y-%m-%d'), dt.strftime('%I:%M %p').lstrip('0'),
            team_names.get(h_id,''), DIVISIONS.get(h_id,'?'),
            '—', '—', team_names.get(v_id,''),
            f"{dp_pred:+.1f}" if dp_pred is not None else '—',
            pred_txt, '—']
    for ci, val in enumerate(vals, 1):
        ref = f'{get_column_letter(ci)}{row}'
        data_style(ws6, ref, val, center=(ci not in (4,8,10)), bg='F0FFF0')
        if ci == 10:
            ws6[ref].fill = PatternFill('solid', fgColor=pbg)
            ws6[ref].font = Font(name='Arial', bold=True, color=pfg, size=10)
    ws6.row_dimensions[row].height = 16
    row += 1

row += 1  # spacer

# ── Section 3: May TBD Games ──────────────────────────────────────────
dp_hdr_row(ws6, row, "🔲  MAY GAMES  —  Weeks 7, 8 & 9  (matchups not yet published)", bg='555555')
row += 1
dp_col_headers(ws6, row); row += 1

for tbd_date, tbd_wk in [("2026-05-10","Week 7"),("2026-05-17","Week 8"),("2026-05-31","Week 9 — Final")]:
    dp_game_num += 1
    dt_tbd = datetime.fromisoformat(tbd_date)
    vals = [dp_game_num, dt_tbd.strftime('%Y-%m-%d'), 'TBD',
            'TBD', '?', '—', '—', 'TBD', '—',
            f"📅 {tbd_wk} — Matchup Not Published", '—']
    for ci, val in enumerate(vals, 1):
        ref = f'{get_column_letter(ci)}{row}'
        c = ws6[ref]
        c.value = val
        c.font = Font(name='Arial', color='888888', size=10, italic=True)
        c.fill = PatternFill('solid', fgColor='F0F0F0')
        c.alignment = Alignment(horizontal='center' if ci not in (4,8,10) else 'left', vertical='center')
        c.border = make_border('thin')
    ws6.row_dimensions[row].height = 16
    row += 1

row += 1  # spacer

# ── Section 4: Season Outlook ─────────────────────────────────────────
dp_hdr_row(ws6, row, "🏒  SEASON OUTLOOK  —  9 Games Total"); row += 1

games_played  = dp_s['GP']
games_remaining = 9 - games_played
cur_pts = dp_s['pts']
w, l, t = dp_s['W'], dp_s['L'], dp_s['T']

scenarios = [
    ("🏆 Best Case",   f"{w+games_remaining}-{l}-{t}",    cur_pts + games_remaining*2,   "Win all remaining"),
    ("⭐ Most Likely",  f"{w+games_remaining}-{l}-{t}",    cur_pts + games_remaining*2,   "Current form → perfect season"),
    ("➡️ Avg. Pace",   f"{w+int(games_remaining*0.7)}-{l+int(games_remaining*0.3)}-{t}", cur_pts+int(games_remaining*0.7)*2, "Win ~70% of remaining"),
    ("⚠️ Downside",    f"{w+int(games_remaining*0.4)}-{l+int(games_remaining*0.6)}-{t}", cur_pts+int(games_remaining*0.4)*2, "Win ~40% of remaining"),
]
ws6.merge_cells(f'A{row}:B{row}')
ws6.merge_cells(f'C{row}:D{row}')
ws6.merge_cells(f'E{row}:F{row}')
ws6.merge_cells(f'G{row}:K{row}')
for col_l, lbl in [('A','Scenario'),('C','Final Record'),('E','Proj. Points'),('G','Notes')]:
    hdr_style(ws6, f'{col_l}{row}', lbl, bg='2E7D32', fg='FFFFFF')
ws6.row_dimensions[row].height = 18
row += 1

for sc_name, sc_rec, sc_pts, sc_note in scenarios:
    ws6.merge_cells(f'A{row}:B{row}')
    ws6.merge_cells(f'C{row}:D{row}')
    ws6.merge_cells(f'E{row}:F{row}')
    ws6.merge_cells(f'G{row}:K{row}')
    for col_l, val in [('A',sc_name),('C',sc_rec),('E',f"{sc_pts} pts"),('G',sc_note)]:
        c = ws6[f'{col_l}{row}']
        c.value = val
        c.font = Font(name='Arial', bold=(col_l=='A'), size=10, color=DP_DARK)
        c.fill = PatternFill('solid', fgColor=DP_BG)
        c.alignment = Alignment(horizontal='left', vertical='center')
        c.border = make_border('thin')
    ws6.row_dimensions[row].height = 16
    row += 1

# Move Disco Pickles to front
wb.move_sheet("🥒 Disco Pickles", offset=-(len(wb.sheetnames)-1))

# ─── Tab colors ──────────────────────────────────────────────────────
ws1.sheet_properties.tabColor = "1F4E79"
ws2.sheet_properties.tabColor = "2E75B6"
ws3.sheet_properties.tabColor = "70AD47"
ws4.sheet_properties.tabColor = "ED7D31"
ws5.sheet_properties.tabColor = "7030A0"
ws6.sheet_properties.tabColor = "1D7A1D"

# ─── Save ─────────────────────────────────────────────────────────────
OUTPUT = '/Users/wgibbons/Desktop/10U_ADV_League_6130.xlsx'
wb.save(OUTPUT)
print(f"Saved: {OUTPUT}")

